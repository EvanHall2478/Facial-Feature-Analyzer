import requests
import openpyxl
import pandas as pd
import json
from PIL import Image
from io import BytesIO

wb = openpyxl.load_workbook(r'images_database.xlsx')
# wb = openpyxl.load_workbook(r"C:\Users\evanh\Software\UofTHacks\NostalgicMobilePhotoAlbum\filterImages\images_database.xlsx")
ws = wb.active

def resize_and_upload_image(file_path):
    # Resize the image
    with Image.open(file_path) as img:
        img_resized = img.resize((150, 150), Image.LANCZOS)
        buffer = BytesIO()
        img_resized.save(buffer, format='JPEG')
        buffer.seek(0)
    
    # Upload the image
    files = {'file': ('resized_image.jpg', buffer, 'image/jpeg')}
    headers = {"X-Cybozu-API-Token": r"6I59Yzq0u6g2L3gc6oQchsMccyfBjpif7e4tfZmx"}
    file_endpoint = f'https://myphotoalbumdb.kintone.com/k/v1/file.json'

    response = requests.post(file_endpoint, headers=headers, files=files)
    if response.status_code == 200:
        file_key = response.json().get('fileKey')
        return file_key
    else:
        print(f'Failed to upload image: {response.text}')
        return None

def addDatabase():
    iterations = []
    rows_list = []
    for index, i in enumerate(ws.iter_rows(values_only=True)): # API can only do 100 entries at a time
        row_data = {'Path': i[0], 'Time': str(i[1]), 'Location': i[2], 'Emotion': i[3]}
        rows_list.append(row_data)
        if index % 99 == 0:
            iterations.append(rows_list)
            rows_list = []
            
    for rows_list in iterations:
        df = pd.DataFrame(rows_list, columns=['Path', 'Time', 'Location', 'Emotion'])
        print(df)

        # Import the necessary libraries
        # Define the Kintone API endpoint
        endpoint = r"https://myphotoalbumdb.kintone.com/k/v1/records.json?app=1&id=1"

        # Define the API key
        api_key = r"6I59Yzq0u6g2L3gc6oQchsMccyfBjpif7e4tfZmx"

        # Prepare the data to be imported
        data = {
            "app": 1,  # Replace with your Kintone app ID # MINE is 1
            "records": []
        }
        for index, row in df.iterrows():
            file_key = resize_and_upload_image(row["Path"])
            record = {
                "image_path": {"value": row["Path"]},
                "date": {"value": row["Time"]},  
                "location": {"value": row["Location"]},
                "emotion": {"value": row["Emotion"]},
                "image": {  
                    "value": [
                        {"fileKey": file_key}
                    ]
                }
            }
            data["records"].append(record)

        # Make the POST request to import the data
        response = requests.post(endpoint, json=data, headers={"X-Cybozu-API-Token": api_key})

        # Check the response status
        if response.status_code == 200:
            print("Data imported successfully!")
        else:
            print("Failed to import data:", response.text)

#Delete batch of 100 records at a time
def deleteDatabase():
    endpoint = r"https://myphotoalbumdb.kintone.com/k/v1/records.json?app=1&id=1"
    api_key = r"6I59Yzq0u6g2L3gc6oQchsMccyfBjpif7e4tfZmx"

    # Get all records from the database
    response = requests.get(endpoint, headers={"X-Cybozu-API-Token": api_key})
    if response.status_code == 200:
        records = response.json().get('records', [])
    else:
        print('Failed to retrieve records:', response.text)
        
    # Prepare the data to be deleted
    data = {
        'app': 1,
        'ids': [record['$id']['value'] for record in records]
    }

    # Delete the records in batches
    batch_size = 100
    for i in range(0, len(data['ids']), batch_size):
        batch_ids = data['ids'][i:i+batch_size]
        batch_data = {
            'app': 1,
            'ids': batch_ids
        }
        response = requests.delete(endpoint, json=batch_data, headers={"X-Cybozu-API-Token": api_key})
        if response.status_code == 200:
            print(f'Deleted {len(batch_ids)} records successfully')
        else:
            print('Failed to delete records:', response.text)
    if response.status_code == 200:
        print('Records deleted successfully')
    else:
        print('Failed to delete records:', response.text)


if __name__ == '__main__': 
    # addDatabase()
    while deleteDatabase():
        deleteDatabase()